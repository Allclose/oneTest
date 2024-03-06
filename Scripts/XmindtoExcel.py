#!/usr/bin/env python
# -*- coding: utf-8 -*-
from xmindparser import xmind_to_dict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment


def write_sub_note_to_excel(datadict, cur='', excel=None):
    '''
    把每个最终节点，带路径写入excel
    :param datadict:
    :param cur:
    :param excel:
    :return:
    '''
    for k, v in datadict.items():
        if type(v) is dict:
            write_sub_note_to_excel(v, cur, excel)
        elif isinstance(v, list):
            for i in v:
                if isinstance(i, dict):
                    write_sub_note_to_excel(i, cur, excel)
        else:
            if not 'topics' in datadict and k == 'title':
                wb = load_workbook(excel)
                ws = wb.active
                ws.append((cur + '-' + v).split('-')[1:])
                wb.save(excel)
                print("{0} : {1}".format(k, cur + '-' + v))
            cur = cur + '-' + v


def format_excel(excelfile):
    '''
    调整excel格式，做合并单元格
    :param excelfile:
    :return:
    '''
    wb = load_workbook(excelfile)
    ws = wb.active
    for cols in ws.iter_cols(max_row=ws.max_row, max_col=ws.max_column):
        srow = cols[0]
        for col in cols[1:]:
            if col.value == srow.value:
                if col == cols[ws.max_row - 1] and col.value is not None:
                    ws.merge_cells(':'.join([srow.coordinate, col.coordinate]))
                    srow.alignment = Alignment(horizontal="center", vertical="center")
            else:
                if srow.value is not None:
                    # print(':'.join([srow.coordinate, col.column_letter + str(col.row - 1)]))
                    ws.merge_cells(':'.join([srow.coordinate, col.column_letter + str(col.row - 1)]))
                    srow.alignment = Alignment(horizontal="center", vertical="center")
                srow = col
    wb.save(excelfile)


if __name__ == '__main__':
    import sys
    xmind_file = r"C:\Users\a\Desktop\思维导图.xmind"  # 输入D:\XX测试用例.xmind
    xmind_dict = xmind_to_dict(xmind_file)
    print(xmind_dict)
    # 生成的excel文件名
    wb = Workbook()
    excelfile = xmind_file.replace('.xmind', '.xlsx')
    wb.save(excelfile)
    wb = load_workbook(excelfile)
    ws = wb.active
    # xmind解析后的dict格式如下：
    # [{'title': '画布 1', 'topic': {'title': 'XX测试用例', 'topics': []}}]
    # 所以需要忽略到第一层的 '画布 1'，直接从topic里内容开始转
    top_topic = xmind_dict[0]['topic']
    write_sub_note_to_excel(top_topic, excel=excelfile)
    # 需要格式二：枢纽布局来逐条查阅的小伙本，直接把下面这行format_excel注释掉即可
    format_excel(excelfile)
